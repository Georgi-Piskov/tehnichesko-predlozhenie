"""
Календарен график за MS Project — ОП2: Кухненско обзавеждане ДГ Чепеларе
Генерира Excel файл съвместим с MS Project.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ============================================================================
# SCHEDULE DATA
# (indent, name, unit, qty, duration, predecessors_1based, resources)
# indent: 0 = summary task (етап), 1 = work task
# duration: 0 за summary (автоматично от деца)
# predecessors: [] или [task_id, ...] — 1-based
# ============================================================================

schedule_data = [
    # ЕТАП 1: ПРОЕКТИРАНЕ И ПОДГОТОВКА (Дни 1-8)
    (0, "ЕТАП 1: ПРОЕКТИРАНЕ И ПОДГОТОВКА", "", "", 0, [], ""),
    (1, "Замерване на помещения - база Здравец и база Елхица", "к-т", 2, 2, [], "Главен проектант; Техник-замерчик"),
    (1, "Изготвяне на работни чертежи - инокс артикули", "к-т", 1, 3, [2], "Главен проектант; CAD оператор"),
    (1, "Изготвяне на работни чертежи - ПДЧ маса и метални шкафчета", "к-т", 1, 2, [2], "Главен проектант; CAD оператор"),
    (1, "Съгласуване на проектните решения с възложителя", "к-т", 1, 3, [3, 4], "Главен проектант; Ръководител проект"),

    # ЕТАП 2: ДОСТАВКА НА МАТЕРИАЛИ (Дни 5-10)
    (0, "ЕТАП 2: ДОСТАВКА НА МАТЕРИАЛИ", "", "", 0, [], ""),
    (1, "Доставка на инокс листове AISI 304", "к-т", 1, 3, [3], "Снабдител"),
    (1, "Доставка на ПДЧ, ABS кант, лепила", "к-т", 1, 2, [4], "Снабдител"),
    (1, "Доставка на стомана DC01, прахова боя", "к-т", 1, 2, [4], "Снабдител"),
    (1, "Доставка на обков, фурнитура, ВиК части", "к-т", 1, 2, [5], "Снабдител"),

    # ЕТАП 3: ПРОИЗВОДСТВО — КОТА -3.00/КУХНЕНСКИ БЛОК
    (0, "ЕТАП 3: ПРОИЗВОДСТВО — КОТА -3.00/КУХНЕНСКИ БЛОК", "", "", 0, [], ""),

    # 1.КУХНЯ
    (0, "1.КУХНЯ", "", "", 0, [], ""),
    (1, "Работен плот с подплот с мивка, инокс, 180/60/85; размер на корито: 80 см. дълж/50 см. ширина/30 см. дълбочина", "бр.", 1, 3, [7, 10], "Специалист инокс; Заварчик TIG"),
    (1, "Шкаф с работен плот с междинен рафт, инокс с отварящи се врати", "бр.", 4, 4, [7, 10], "Специалист инокс; Заварчик TIG"),
    (1, "Стенен шкаф с междинен рафт, инокс с отварящи се врати", "бр.", 1, 2, [14], "Специалист инокс"),
    (1, "Работна маса с подплот, инокс", "бр.", 2, 2, [14], "Специалист инокс"),

    # 2.СЕРВИЗНО ПОМЕЩЕНИЕ
    (0, "2.СЕРВИЗНО ПОМЕЩЕНИЕ", "", "", 0, [], ""),
    (1, "Работен плот с подплот, инокс", "бр.", 2, 2, [15, 16], "Специалист инокс"),

    # 3.СТАЯ ЗА ПОЧИВКА
    (0, "3.СТАЯ ЗА ПОЧИВКА", "", "", 0, [], ""),
    (1, "Маса кръгла- плот от ПДЧ и крака от дърво масив с влагоустойчиво покритие на плота и краката", "бр.", 1, 2, [8], "Дървообработчик; Лакировчик"),
    (1, "Стол с метална основа с кожена тапицерия", "бр.", 4, 2, [9], "Метален конструктор; Тапицер"),

    # 4.МИЯЛНО ПОМЕЩЕНИЕ
    (0, "4.МИЯЛНО ПОМЕЩЕНИЕ", "", "", 0, [], ""),
    (1, "Работен плот с две мивки 60х50 с душ батерии, инокс, размерите на коритото - ш/дълж/дълб. 50см./80см./30см", "бр.", 1, 3, [18], "Специалист инокс; Заварчик TIG"),
    (1, "Мивка дълбока 90 / 50 с душ батерия, инокс", "бр.", 1, 2, [23], "Специалист инокс"),
    (1, "Стелаж за чиста посуда, инокс, с 3 рафта", "бр.", 1, 2, [23], "Специалист инокс"),

    # 5.ПОМЕЩЕНИЕ ЗА ПАСИРАНЕ НА ХРАНА
    (0, "5.ПОМЕЩЕНИЕ ЗА ПАСИРАНЕ НА ХРАНА", "", "", 0, [], ""),
    (1, "Работен плот с подплот с мивка 50х40, инокс", "бр.", 1, 2, [24, 25], "Специалист инокс; Заварчик TIG"),
    (1, "Шкаф с работен плот и междинен рафт, инокс, с отварящи се врати", "бр.", 1, 2, [27], "Специалист инокс"),

    # 6.ПОМЕЩЕНИЕ ПОДГОТВИТЕЛНО
    (0, "6.ПОМЕЩЕНИЕ ПОДГОТВИТЕЛНО", "", "", 0, [], ""),
    (1, "Работен плот с подплот с две мивки 50х40, инокс", "бр.", 1, 2, [28], "Специалист инокс; Заварчик TIG"),
    (1, "Шкаф с плот и междинен рафт, инокс, с отварящи се врати", "бр.", 1, 2, [30], "Специалист инокс"),

    # 7.СКЛАД ПОЧИСТВАЩ ИНВЕНТАР, ХИМИКАЛИ
    (0, "7.СКЛАД ПОЧИСТВАЩ ИНВЕНТАР, ХИМИКАЛИ", "", "", 0, [], ""),
    (1, "Аусгус мивка с диспенсър за течен сапун, инокс", "бр.", 1, 2, [31], "Специалист инокс"),
    (1, "Шкаф за химикали и биоциди, инокс с четири рафта, с възможност за заключване", "бр.", 1, 2, [33], "Специалист инокс"),

    # 8.СЕРВИЗНО ПОМЕЩЕНИЕ
    (0, "8.СЕРВИЗНО ПОМЕЩЕНИЕ (кота -3.00)", "", "", 0, [], ""),
    (1, "Работен плот с подплот, инокс", "бр.", 2, 2, [34], "Специалист инокс"),

    # 9.ПОДГОТВИТЕЛНО ПОМЕЩЕНИЕ
    (0, "9.ПОДГОТВИТЕЛНО ПОМЕЩЕНИЕ", "", "", 0, [], ""),
    (1, "Работен плот с подплот с две мивки 50х40, инокс", "бр.", 1, 2, [36], "Специалист инокс; Заварчик TIG"),
    (1, "Шкаф с междинен рафт с плъзгащи врати, инокс", "бр.", 1, 2, [38], "Специалист инокс"),

    # 13.СКЛАД-ЗЕЛЕНЧУЦИ/ПЛОДОВЕ
    (0, "13.СКЛАД-ЗЕЛЕНЧУЦИ/ПЛОДОВЕ", "", "", 0, [], ""),
    (1, "Складов стелаж, инокс с пет рафта", "бр.", 3, 3, [39], "Специалист инокс"),

    # 14.СКЛАД ПАКЕТАЖ
    (0, "14.СКЛАД ПАКЕТАЖ", "", "", 0, [], ""),
    (1, "Складов стелаж, инокс с пет рафта", "бр.", 3, 3, [41], "Специалист инокс"),

    # 15.СЪБЛЕКАЛНЯ ПЕРСОНАЛ
    (0, "15.СЪБЛЕКАЛНЯ ПЕРСОНАЛ", "", "", 0, [], ""),
    (1, "Шкафчета за облекло на персонала метален", "бр.", 4, 3, [9, 21], "Метален конструктор; Бояджия"),

    # ЕТАП 4: ПРОИЗВОДСТВО — КОТА +3,10 / КОТА +0,00
    (0, "ЕТАП 4: ПРОИЗВОДСТВО — СЕРВИЗНИ ПОМЕЩЕНИЯ ГОРНИ КОТИ", "", "", 0, [], ""),

    # кота +3,10 СЕКЦИЯ 1
    (0, "кота +3,10 СЕКЦИЯ 1 — Сервизно помещение за храна", "", "", 0, [], ""),
    (1, "Шкаф с плот и мивка 40х50, с депо за отпадъци, инокс", "бр.", 1, 2, [43], "Специалист инокс"),
    (1, "Кош за отпадъци, инокс", "бр.", 1, 1, [49], "Специалист инокс"),

    # кота +3,10 СЕКЦИЯ 2
    (0, "кота +3,10 СЕКЦИЯ 2 — Сервизно помещение за храна", "", "", 0, [], ""),
    (1, "Шкаф с плот и мивка 40х50, с депо за отпадъци, инокс", "бр.", 1, 2, [50], "Специалист инокс"),
    (1, "Кош за отпадъци, инокс", "бр.", 1, 1, [52], "Специалист инокс"),

    # кота +0,00 СЕКЦИЯ 1
    (0, "кота +0,00 СЕКЦИЯ 1 — Сервизно помещение за храна", "", "", 0, [], ""),
    (1, "Шкаф с плот и мивка 40х50, с депо за отпадъци, инокс", "бр.", 1, 2, [53], "Специалист инокс"),
    (1, "Кош за отпадъци, инокс", "бр.", 1, 1, [55], "Специалист инокс"),

    # кота +0,00 СЕКЦИЯ 2
    (0, "кота +0,00 СЕКЦИЯ 2 — Сервизно помещение за храна", "", "", 0, [], ""),
    (1, "Шкаф с плот и мивка 40х50, с депо за отпадъци, инокс", "бр.", 1, 2, [56], "Специалист инокс"),
    (1, "Кош за отпадъци, инокс", "бр.", 1, 1, [58], "Специалист инокс"),

    # ЕТАП 5: КОНТРОЛ НА КАЧЕСТВОТО
    (0, "ЕТАП 5: КОНТРОЛ НА КАЧЕСТВОТО", "", "", 0, [], ""),
    (1, "Контрол на размерите и качеството на инокс артикулите", "к-т", 1, 2, [59], "Контрольор качество"),
    (1, "Тест за корозионна устойчивост (3% солен разтвор, 24ч)", "к-т", 1, 1, [61], "Контрольор качество"),
    (1, "Контрол на ПДЧ, метални шкафчета, тапицерия", "к-т", 1, 1, [45, 61], "Контрольор качество"),
    (1, "Документиране — сертификати, протоколи от изпитвания", "к-т", 1, 1, [62, 63], "Контрольор качество"),

    # ЕТАП 6: ОПАКОВАНЕ И ДОСТАВКА
    (0, "ЕТАП 6: ОПАКОВАНЕ И ДОСТАВКА", "", "", 0, [], ""),
    (1, "Опаковане на артикулите — защитно фолио, пяна, палети", "к-т", 1, 2, [64], "Опаковчици"),
    (1, "Транспорт до база Здравец (ул. Здравец №3)", "к-т", 1, 1, [66], "Шофьор-експедитор; Товарач"),
    (1, "Транспорт до база Елхица (ул. Перелик №1)", "к-т", 1, 1, [67], "Шофьор-експедитор; Товарач"),
    (1, "Приемане и складиране на обекта", "к-т", 1, 1, [67, 68], "Ръководител проект; Склад. работник"),

    # ЕТАП 7: МОНТАЖ
    (0, "ЕТАП 7: МОНТАЖ", "", "", 0, [], ""),
    (1, "Монтаж кота -3.00 — работни плотове с мивки, ВиК свързване", "к-т", 1, 3, [69], "Ръководител монтаж; Монтажник; ВиК"),
    (1, "Монтаж кота -3.00 — шкафове, стелажи, стенни елементи", "к-т", 1, 3, [71], "Ръководител монтаж; Монтажник"),
    (1, "Монтаж кота -3.00 — маса, столове, шкафчета за облекло", "к-т", 1, 2, [71], "Монтажник"),
    (1, "Монтаж кота +3,10 и кота +0,00 — шкафове с мивки, кошове", "к-т", 1, 2, [72], "Монтажник; ВиК"),

    # ЕТАП 8: ПРИЕМАНЕ И ДОКУМЕНТАЦИЯ
    (0, "ЕТАП 8: ПРИЕМАНЕ И ДОКУМЕНТАЦИЯ", "", "", 0, [], ""),
    (1, "Финална проверка и тестване на монтираното оборудване", "к-т", 1, 1, [72, 73, 74], "Контрольор качество; Ръководител проект"),
    (1, "Подписване на приемо-предавателен протокол", "к-т", 1, 1, [76], "Ръководител проект"),
    (1, "Предаване на гаранционна документация", "к-т", 1, 1, [76], "Ръководител проект"),
]

# ============================================================================
# FORWARD-PASS SCHEDULING
# ============================================================================
n = len(schedule_data)
is_summary = [schedule_data[i][0] == 0 for i in range(n)]
duration_arr = [schedule_data[i][4] for i in range(n)]
preds_arr = [schedule_data[i][5] for i in range(n)]

start_day = [0] * n
end_day = [0] * n

# Pass 1: compute work tasks
for i in range(n):
    if is_summary[i]:
        continue
    dur = duration_arr[i]
    preds = preds_arr[i]
    if not preds:
        start_day[i] = 1
    else:
        start_day[i] = max(end_day[p - 1] for p in preds) + 1
    end_day[i] = start_day[i] + dur - 1

# Pass 2: compute summary tasks from children
def find_children(summary_idx):
    """Find all work tasks belonging to this summary."""
    children = []
    level = schedule_data[summary_idx][0]
    for j in range(summary_idx + 1, n):
        if schedule_data[j][0] <= level:
            break
        if not is_summary[j]:
            children.append(j)
    return children

for i in range(n):
    if not is_summary[i]:
        continue
    children = find_children(i)
    if children:
        starts = [start_day[c] for c in children if start_day[c] > 0]
        ends = [end_day[c] for c in children if end_day[c] > 0]
        if starts and ends:
            start_day[i] = min(starts)
            end_day[i] = max(ends)

total_days = max(end_day)
print(f"Общо работни дни (критичен път): {total_days}")

# ============================================================================
# GENERATE EXCEL
# ============================================================================
wb = Workbook()
ws = wb.active
ws.title = "Календарен график"

# Column headers
headers = ["№", "Наименование", "Мярка", "Количество", "Работни дни",
           "Предходни задачи", "Начало", "Край", "Ресурси"]

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
summary_font = Font(bold=True, size=10)
summary_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
normal_font = Font(size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    cell.number_format = '@'

# Write data
for i in range(n):
    row = i + 2
    indent, name, unit, qty, dur, preds, resources = schedule_data[i]

    # Task number
    cell = ws.cell(row=row, column=1, value=str(i + 1))
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    # Name (with indent for subtasks)
    prefix = "    " if indent == 1 else ""
    cell = ws.cell(row=row, column=2, value=prefix + name)
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(wrap_text=True)

    # Unit
    cell = ws.cell(row=row, column=3, value=str(unit))
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    # Quantity
    qty_str = str(qty) if qty != "" else ""
    cell = ws.cell(row=row, column=4, value=qty_str)
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    # Duration
    dur_str = str(dur) if dur > 0 else ""
    cell = ws.cell(row=row, column=5, value=dur_str)
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    # Predecessors
    pred_str = ";".join(str(p) for p in preds) if preds else ""
    cell = ws.cell(row=row, column=6, value=pred_str)
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    # Start
    start_str = str(int(start_day[i])) if start_day[i] > 0 else ""
    cell = ws.cell(row=row, column=7, value=start_str)
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    # End
    end_str = str(int(end_day[i])) if end_day[i] > 0 else ""
    cell = ws.cell(row=row, column=8, value=end_str)
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    # Resources
    cell = ws.cell(row=row, column=9, value=str(resources))
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(wrap_text=True)

    # Apply formatting
    if is_summary[i]:
        for col in range(1, 10):
            ws.cell(row=row, column=col).font = summary_font
            ws.cell(row=row, column=col).fill = summary_fill
    else:
        for col in range(1, 10):
            ws.cell(row=row, column=col).font = normal_font

# Column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 65
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 40

# ============================================================================
# SHEET 2: RESOURCES
# ============================================================================
ws2 = wb.create_sheet("Ресурси")

resources_data = [
    ("Ръководител проект", "Трудов", "Дни 1-45"),
    ("Главен проектант", "Трудов", "Дни 1-8"),
    ("CAD оператор", "Трудов", "Дни 3-7"),
    ("Техник-замерчик", "Трудов", "Дни 1-2"),
    ("Снабдител", "Трудов", "Дни 3-10"),
    ("Специалист инокс", "Трудов", "Дни 8-35"),
    ("Заварчик TIG", "Трудов", "Дни 8-30"),
    ("Дървообработчик", "Трудов", "Дни 10-14"),
    ("Метален конструктор", "Трудов", "Дни 10-18"),
    ("Тапицер", "Трудов", "Дни 12-16"),
    ("Лакировчик", "Трудов", "Дни 10-14"),
    ("Бояджия (прахово)", "Трудов", "Дни 14-18"),
    ("Контрольор качество", "Трудов", "Дни 35-40"),
    ("Опаковчици", "Трудов", "Дни 39-40"),
    ("Шофьор-експедитор", "Трудов", "Дни 41-42"),
    ("Товарач", "Трудов", "Дни 41-42"),
    ("Ръководител монтаж", "Трудов", "Дни 43-45"),
    ("Монтажник", "Трудов", "Дни 43-45"),
    ("ВиК специалист", "Трудов", "Дни 43-45"),
    ("Складов работник", "Трудов", "Дни 42-43"),
    ("Инокс листове AISI 304", "Материален", "Дни 5-7"),
    ("ПДЧ влагоустойчив Е1", "Материален", "Дни 5-6"),
    ("Стомана DC01", "Материален", "Дни 5-6"),
    ("Обков, фурнитура, ВиК части", "Материален", "Дни 8-9"),
    ("CNC лазерна машина", "Оборудване", "Дни 8-35"),
    ("TIG заваръчен апарат", "Оборудване", "Дни 8-30"),
    ("Хидравлична преса", "Оборудване", "Дни 8-35"),
    ("Камион Mercedes Atego", "Оборудване", "Дни 41-42"),
]

res_headers = ["№", "Ресурс", "Тип", "Период на ангажиране"]
for col, h in enumerate(res_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    cell.number_format = '@'

for idx, (res_name, res_type, period) in enumerate(resources_data, 1):
    row = idx + 1
    cell = ws2.cell(row=row, column=1, value=str(idx))
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    cell = ws2.cell(row=row, column=2, value=res_name)
    cell.number_format = '@'
    cell.border = thin_border

    cell = ws2.cell(row=row, column=3, value=res_type)
    cell.number_format = '@'
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    cell = ws2.cell(row=row, column=4, value=period)
    cell.number_format = '@'
    cell.border = thin_border

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 25

# Save
output_file = "GRAFIK DG_KUHNQ/Календарен_график_ОП2_кухня.xlsx"
wb.save(output_file)
print(f"Файлът е записан: {output_file}")
print(f"Брой задачи: {n}")
print(f"Критичен път: {total_days} работни дни")
